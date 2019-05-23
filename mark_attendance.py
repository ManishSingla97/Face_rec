import sqlite3

from openpyxl import Workbook, load_workbook

from openpyxl.utils import get_column_letter, column_index_from_string

from openpyxl.cell import Cell

import time
#get current date

wb = load_workbook(filename = "reports.xlsx")
	
sheet = wb['CSA']
currentDate = time.strftime("%d_%m_%y")

def getDateColumn():
	
	for i in range(1, len(list(sheet.rows)[0]) + 1):
		
		col = get_column_letter(i)
		
		if sheet['%s%s'% (col,'1')].value == currentDate:
		
			return col

def attendance(Id):
	
	
	
	



	connect = sqlite3.connect("mydatabase.db")
#c = connect.cursor()

	
	attend = [0 for i in range(60)]	
	
	personId =  Id
	
	cur = connect.execute("SELECT * FROM Students WHERE regNo = (?)", (personId,))
				#print("cur = {}".format(cur))
	
	for row in cur:
		
		#print("hello")
		
		#print("row = {}".format(row))
		
		attend[int(row[2])] += 1
		
		print("---------- " + row[0] + " recognized ----------")
		
		time.sleep(6)
		
	for row in range(2, len(list(sheet.columns)[0]) + 1):
		
		rn = sheet.cell(row = row, column  =1).value
		
		print(rn)
		
		if rn is not None:
		
			#print("rn = {}".format(rn))
		
			#rn = row
		
			if attend[int(rn)] != 0:
		
				col = getDateColumn()
		
				#print("col = {}".format(col))
		
				sheet['%s%s' % (col, str(row))] = 1

	wb.save(filename = "reports.xlsx")	 	
